#!/usr/bin/python

import xlsxwriter
import xlwt
import sys
import openpyxl
from openpyxl.cell import get_column_letter
import numpy as np

def panel2excel(datadir, filename, header, panelcells, metadata):
    wb = xlwt.Workbook(encoding='utf')

    f_short_name = "Data"
    ws = wb.add_sheet(str(f_short_name))
    style_string = "font: bold on; borders: bottom medium"
    style = xlwt.easyxf(style_string)

    i = 0
    col_width = 256 * 20
    # font
    font = xlwt.Font()
    font.bold = True
    style.font = font
    for row in header:
        ws.col(i).width = col_width
        ws.write(0, i, row, style=style)
        i = i+1

    f_copyrights_name = "Copyrights"
    copyrights = "Clio Infra (2015) http://www.clio-infra.eu"

    wscop = wb.add_sheet(str(f_copyrights_name))
    col_width = 256 * 30
    for i in range(0,4):
        wscop.col(i).width = col_width 
    wscop.write(0,0,copyrights)

    lineID = 0
    # If metadata exist
    for item in metadata:
	if item['url']:
	    lineID = lineID + 1
	    wscop.write(lineID,0, str(item['url']))
	    #wscop.cell(lineID,0).hyperlink = str(item['url'])
	    wscop.write(lineID,1, str(item['name']))
	    wscop.write(lineID,2, str(item['description']))
	    wscop.write(lineID,3, str(item['citation']))

    i = 0
    for dataitem in panelcells:
        i = i + 1
	dataonly = []
	# Append country and years
  	dataonly.append(dataitem[0])
        dataonly.append(dataitem[1])
        # Skip all country codes
	for j in range(2, len(dataitem)):
	    #if (j/2)*2 == j:
	    if j:
	        dataonly.append(dataitem[j])

	dataitem = dataonly
        for j in range(0, len(dataitem)):
            value = dataitem[j]
            if value == 'NaN':
                value = ''
	    if value > -1000000000000:
                ws.write(i, j, value)

    fullpath = datadir + "/" + filename
    wb.save(fullpath)

    return fullpath 

def individual_dataset(datadir, filename, indicator, units, inputdatahub, data, codes, metadata):
    wb = xlwt.Workbook(encoding='utf')

    f_short_name = "Data"
    ws = wb.add_sheet(str(f_short_name))
    style_string = "font: bold on; borders: bottom medium"
    style = xlwt.easyxf(style_string)
    simplestyle = xlwt.easyxf(style_string)

    i = 0
    col_width = 256 * 20
    # font
    font = xlwt.Font()
    font.bold = True
    style.font = font
    font = xlwt.Font()
    font.bold = False
    simplestyle.font = font

    (x,y) = (0,0)
    startX = 2
    startY = 2
    x = startX    
    ws.col(0).width = col_width
    ws.col(1).width = col_width + 256 * 40
    ws.write(0, 0, indicator, style=style)
    ws.write(1, 0, units, style=style)
    ws.write(2, 0, 'Code', style=style)
    ws.write(2, 1, 'Continent, Region, Country', style=style)

    datahub = {}
    limit = 224
    limy = 0
    for year in sorted(inputdatahub):
        try:
            dataitem = data[year]
        except:
            dataitem = []

        for item in dataitem:
            datacode = item[1]
            datavalue = item[6]
	    if datavalue == 'NaN':
		datavalue = ''

	    if datavalue:
	        datahub[year] = inputdatahub[year]
	  	#print str(year) + ' ' + str(datavalue)
	    limy = limy + 1
    
    codeID = startX 
    for code in sorted(codes):
        codeID = codeID + 1
        country = codes[code]
        ws.write(codeID, 0, code)
        ws.write(codeID, 1, country)
    
    y = startY 
    for year in sorted(datahub):        
        ws.write(2, y, year, style=style)
        y = y + 1
    
    y = startY
    for year in sorted(datahub):
        codes = datahub[year]
	dataitem = []
	try:
            dataitem = data[year]
	except:
	    dataitem = []
        #print str(year) + str(codes)
        print dataitem
        dataindex = {}
        for item in dataitem:
            datacode = item[1]
            datavalue = item[6]                        
            if datavalue == 'NaN':
                datavalue = ''
            dataindex[str(datacode)] = datavalue                
        
        # Filling columns
        for code in sorted(codes):
            country = codes[code]               
            x = x + 1
	    try:
                value = dataindex[str(code)]
	    except:
		value = ''
            print str(x) + ',' + str(y) + ':' + str(value)
            ws.write(x, y, value)
        
        # New row
        x = startX 
        y = y + 1
    
    fullpath = datadir + "/" + filename
    wb.save(fullpath)    
    return fullpath

def create_excel_dataset(fullpath, geocoder, metadata, metacolumns, yearscolumns, dataset, datayears, datactr):
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Data"
    countryinfo = []

    col_width = 256 * 20
    # Forming header
    order = []
    start = 2
    c = ws.cell(row=0, column=0)
    ws.column_dimensions[get_column_letter(1)].width = 20
    c.value = metadata['title']
    c.style.font.bold = True
    c = ws.cell(row=1, column=0)
    c.value = metadata['units']
    c.style.font.bold = True
    i = 0
    for newcolumn in metacolumns:
        c = ws.cell(row=start, column=i)
        c.value = newcolumn
        c.style.font.bold = True
        order.append(newcolumn)
        ws.column_dimensions[get_column_letter(i+1)].width = 15
        i = i + 1
    for year in sorted(yearscolumns):
        c = ws.cell(row=start, column=i)
        c.value = str(year)
        c.style.font.bold = True
        i = i + 1

    i = start
    for idc in geocoder.index:
        i = i + 1
        #metadata = geocoder.ix[int(idc)]
        j = 0        
        for columnname in metacolumns:
            thisvalue = ''
            #c = ws.cell(row=i, column=j)            
            try:
                thisvalue = str(geocoder.ix[int(idc)][columnname])                
            except:
                skip = 1
                
	    c = ws.cell(row=i, column=j)
            if thisvalue != '':
                c.value = thisvalue
	    else:
		c.value = ''
            j = j + 1

	# Searching for active countries
        activectr = ''
	try:
	    if int(idc) in datactr:
	        activectr = 'yes'
	except:
	    activectr = ''

	if activectr == 'yes':
	    # Checking active years
	    for year in yearscolumns:
                if year in datayears:
                    try:
			tmpval = dataset.ix[int(idc)][int(year)]
                        c = ws.cell(row=i, column=j)
                        c.value = tmpval
                    except:
                        skip = 'on'                
                        #c.value = randint(0,9)
                j = j + 1

    wb.save(fullpath)
   
    return fullpath
