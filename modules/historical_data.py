#!/usr/bin/python
from pandas.io.json import json_normalize
import pandas as pd
import numpy as np
import re
import os
import sys
import openpyxl
from random import randint
from openpyxl.cell import get_column_letter
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../modules')))
from search import dataset_search, getindicators, loadjson

def dataframe_filter(codes, maincode, clioframe):    
    result = {}
    navicolumns = {}
    c = clioframe.columns.tolist()
    dataID = 0
    lastdata = []
    ycodes = []
        
    allcolumns = []
    if not codes:
        for cid in clioframe[maincode]:
            try:
                cid = int(cid)
                allcolumns.append(cid)
            except:
                skip = 1
        codes = allcolumns
        
    for code in codes:
        datafilter = clioframe[clioframe[maincode] == code]
        data = datafilter
        #.transpose()
        c = datafilter.columns.tolist()
        values = datafilter.get_values()
        # Delete extra columns
        for item in data:
            try:
                #print item
                year = int(item)
            except:
                del data[item]
                navicolumns[item] = values
        if dataID == 0:
            data.reset_index(col_fill=[code])
            totaldf = data
            dataID = dataID + 1
            ycodes.append(int(code))
        else:        
            data.reset_index(col_fill=[code])
            lastdata = data
            totaldf = pd.merge(totaldf, data, how='outer')
            ycodes.append(int(code))
        
        result[code] = data
        
    empty = {}
    for year in lastdata:
        empty[year] = ''
    
    return (totaldf, navicolumns, result, ycodes, empty)

def get_info(clioframe):
    header = clioframe.columns
    countryinfo = {}    
    yearscolumns = {}
    line = []
    geo = {}
    for col in header:
        try:
            year = int(col)
            yearscolumns[year] = year
        except:
            countryinfo[col] = clioframe[col]
            line = clioframe[col].values
               
    vocab = {}
    for geoid in clioframe.index:        
        geoinfo = []                
        geo[geoid] = ''
        lex = {}
        for colname in sorted(countryinfo):            
            sgeoid = str(geoid)            
            item = countryinfo[colname][sgeoid]                                    
            lex[colname] = item
        vocab[sgeoid] = lex
        
    return (countryinfo, line, vocab, yearscolumns)
            
def save_historical_dataset(fullpath, geocoder, countryinfo, yearscolumns, dataset):
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Data"

    col_width = 256 * 20
    # Forming header    
    order = []
    start = 2
    c = ws.cell(row=0, column=0)
    ws.column_dimensions[get_column_letter(1)].width = 20
    c.value = 'Dataset title'
    c.style.font.bold = True
    c = ws.cell(row=1, column=0)
    c.value = 'Units'
    c.style.font.bold = True
    i = 0
    for newcolumn in sorted(countryinfo):        
        c = ws.cell(row=start, column=i)
        c.value = newcolumn
        c.style.font.bold = True
        order.append(newcolumn)
        ws.column_dimensions[get_column_letter(i+1)].width = 15
        i = i + 1
    for year in sorted(yearscolumns):        
        c = ws.cell(row=start, column=i)
        c.value = str(year)
        c.style.font.bold = True        
        i = i + 1        
        
    i = start
    for idc in geocoder:
        i = i + 1
        metadata = geocoder[idc]
        j = 0
        for columnname in order:            
            c = ws.cell(row=i, column=j)
            c.value = metadata[columnname]
            j = j + 1        
        for year in yearscolumns:            
            c = ws.cell(row=i, column=j)
            try:
                tmpval = dataset[str(idc)][str(year)].values                
                c.value = tmpval[0]                
            except:
                c.value = ''
		c.value = randint(0,9)
            j = j + 1
        
    wb.save(fullpath)
    
    return fullpath

def match_geocoders(maincode, tmpgeocoder, geo, countryinfo, limityear):
    geocoder = {}
    for item in tmpgeocoder:
        cid = item['id']
        name = item['name']
        ctrinfo = item['year']
        g = re.search(r'(\d+)\s+\-\-\s+(\d+)', ctrinfo)    
        try:
            startyear = g.group(1)
            endyear = g.group(2)
        except:
            (startyear, endyear) = ('', '')
    
        geoitem = {}
        geoitem[maincode] = str(cid)
        geoitem['start date'] = int(startyear)
        geoitem['end date'] = int(endyear)
        geoitem['country'] = name
    
        # Filling metadata
        for newcolumn in countryinfo:
            try:            
                geoitem[newcolumn] = str(geo[str(cid)][newcolumn])
            except:            
                geoitem[newcolumn] = ''            
    
	if limityear:
            if int(startyear) < int(limityear):
                geocoder[cid] = geoitem
	else:
	    geocoder[cid] = geoitem
    
        try:
            geoinfra = geo[cid]
        except:        
            skip = 1
    return geocoder

def datasetreader(handle, config):
    coder = {}
    maindata = []
    apiroot = config['apiroot'] + "/api/datasets?handle=" + handle
    datasets = loadjson(apiroot)
    for item in datasets:
        maindata = item['data']
    return maindata

def produce_historicaldata(config, fullfilepath, clioframe, codes, maincode, limityear):
    clioframe.index = clioframe[maincode]
    (countryinfo, line, geoexcel, yearscolumns) = get_info(clioframe)
    # Filter data from dataframe
    (totaldf, navicolumns, result, ycodes, emptyrow) = dataframe_filter(codes, maincode, clioframe)
    # Geocoder
    geocoderapi = config['geocoderapi']
    tmpgeocoder = loadjson(geocoderapi)
    geocoder = match_geocoders(maincode, tmpgeocoder, geoexcel, countryinfo, limityear)
    # Save as Excel file
    datafile = save_historical_dataset(fullfilepath, geocoder, countryinfo, yearscolumns, result)
    return fullfilepath
